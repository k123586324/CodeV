from openpyxl import Workbook


def Conversion_Factor():
    CF_array = [[0] * 14 for i in range(14)]
    # n=0
    CF_array[0][0] = 1
    # n=1
    CF_array[1][0] = -5
    CF_array[1][1] = 6
    # n=2
    CF_array[2][0] = 15
    CF_array[2][1] = -42
    CF_array[2][2] = 28
    # n=3
    CF_array[3][0] = -35
    CF_array[3][1] = 168
    CF_array[3][2] = -252
    CF_array[3][3] = 120
    # n=4
    CF_array[4][0] = 70
    CF_array[4][1] = -504
    CF_array[4][2] = 1260
    CF_array[4][3] = -1320
    CF_array[4][4] = 495
    # n=5
    CF_array[5][0] = -126
    CF_array[5][1] = 1260
    CF_array[5][2] = -4620
    CF_array[5][3] = 7920
    CF_array[5][4] = -6435
    CF_array[5][5] = 2002
    # n=6
    CF_array[6][0] = 210
    CF_array[6][1] = -2772
    CF_array[6][2] = 13860
    CF_array[6][3] = -34320
    CF_array[6][4] = 45045
    CF_array[6][5] = -30030
    CF_array[6][6] = 8008
    # n=7
    CF_array[7][0] = -330
    CF_array[7][1] = 5544
    CF_array[7][2] = -36036
    CF_array[7][3] = 120120
    CF_array[7][4] = -225225
    CF_array[7][5] = 240240
    CF_array[7][6] = -136136
    CF_array[7][7] = 31824
    # n=8
    CF_array[8][0] = 495
    CF_array[8][1] = -10296
    CF_array[8][2] = 84084
    CF_array[8][3] = -360360
    CF_array[8][4] = 900900
    CF_array[8][5] = -1361360
    CF_array[8][6] = 1225224
    CF_array[8][7] = -604656
    CF_array[8][8] = 125970
    # n=9
    CF_array[9][0] = -715
    CF_array[9][1] = 18018
    CF_array[9][2] = -180180
    CF_array[9][3] = 960960
    CF_array[9][4] = -3063060
    CF_array[9][5] = 6126120
    CF_array[9][6] = -7759752
    CF_array[9][7] = 6046560
    CF_array[9][8] = -2645370
    CF_array[9][9] = 497420
    # n=10
    CF_array[10][0] = 1001
    CF_array[10][1] = -30030
    CF_array[10][2] = 360360
    CF_array[10][3] = -2333760
    CF_array[10][4] = 9189180
    CF_array[10][5] = -23279256
    CF_array[10][6] = 38798760
    CF_array[10][7] = -42325920
    CF_array[10][8] = 29099070
    CF_array[10][9] = -11440660
    CF_array[10][10] = 1961256
    # n=11
    CF_array[11][0] = -1365
    CF_array[11][1] = 48048
    CF_array[11][2] = -680680
    CF_array[11][3] = 5250960
    CF_array[11][4] = -24942060
    CF_array[11][5] = 77597520
    CF_array[11][6] = -162954792
    CF_array[11][7] = 232792560
    CF_array[11][8] = -223092870
    CF_array[11][9] = 137287920
    CF_array[11][10] = -49031400
    CF_array[11][11] = 7726160
    # n=12
    CF_array[12][0] = 1820
    CF_array[12][1] = -74256
    CF_array[12][2] = 1225224
    CF_array[12][3] = -11085360
    CF_array[12][4] = 62355150
    CF_array[12][5] = -232792560
    CF_array[12][6] = 597500904
    CF_array[12][7] = -1070845776
    CF_array[12][8] = 1338557220
    CF_array[12][9] = -1144066000
    CF_array[12][10] = 637408200
    CF_array[12][11] = -208606320
    CF_array[12][12] = 30421755
    # n=13
    CF_array[13][0] = -2380
    CF_array[13][1] = 111384
    CF_array[13][2] = -2116296
    CF_array[13][3] = 22170720
    CF_array[13][4] = -145495350
    CF_array[13][5] = 640179540
    CF_array[13][6] = -1963217256
    CF_array[13][7] = 4283383104
    CF_array[13][8] = -6692786100
    CF_array[13][9] = 7436429000
    CF_array[13][10] = -5736673800
    CF_array[13][11] = 2920488480
    CF_array[13][12] = -882230895
    CF_array[13][13] = 119759850
    return CF_array


# print(Conversion_Factor())
text = []
with open('2G3P_FOV50_22_V5_VIG.seq', 'r', ) as f:
    for line in f:
        text.append(line)

lens_data = []
for i in range(0, len(text)):
    # print(text[i])
    if len(text[i].split()) > 1:
        if text[i].split()[0] == "SO":
            lens = [0] * 22
            for l in range(0, len(lens)):
                lens[l] = ""
            lens[0] = "Standard"
            lens[1] = "Object"
            if float(text[i].split()[1]) == 0:
                lens[2] = "Infinity"
            else:
                lens[2] = float(text[i].split()[1])
            if float(text[i].split()[2]) > 10e5:
                lens[3] = "Infinity"
            else:
                lens[3] = float(text[i].split()[2])
            lens_data.append(lens)
        elif text[i].split()[0] == "SI":
            lens = [0] * 22
            for l in range(0, len(lens)):
                lens[l] = ""
            lens[0] = "Standard"
            lens[1] = "Image"
            if float(text[i].split()[1]) == 0:
                lens[2] = "Infinity"
            else:
                lens[2] = float(text[i].split()[1])
            if float(text[i].split()[2]) > 10e5:
                lens[3] = "Infinity"
            else:
                lens[3] = float(text[i].split()[2])
            lens_data.append(lens)
        if text[i].split()[0] == "S":
            lens = [0] * 22
            # Surface Name
            if text[i + 2].split()[0] == "SLB":
                lens[1] = text[i + 2].split()[1]
            else:
                lens[1] = ""
            # Radius
            if float(text[i].split()[1]) == 0:
                lens[2] = "Infinity"
            else:
                lens[2] = float(text[i].split()[1])
            # Thickness
            lens[3] = float(text[i].split()[2])
            # Glass
            if len(text[i].split()) == 4:
                lens[4] = text[i].split()[3]
            else:
                lens[4] = ""

            qtype = 0
            if text[i + 2].split()[1] == "QCN":
                qtype = 2
                lens[0] = "Q-Type"
            elif text[i + 3].split()[1] == "QCN":
                qtype = 3
                lens[0] = "Q-Type"
            else:
                lens[0] = "Standard"

            if qtype > 1:
                # K
                if "SCO K" in text[i + qtype + 1]:
                    lens[5] = float(text[i + qtype + 1].split()[2].split(";")[0])
                else:
                    lens[5] = 0
                    qtype = qtype - 1
                # NRADIUS
                lens[6] = float(text[i + qtype + 2].split()[2].split(";")[0])
                if "C4" in text[i + qtype + 3]:
                    lens[8] = float(text[i + qtype + 3].split()[2].split(";")[0])
                if "C5" in text[i + qtype + 4]:
                    lens[9] = float(text[i + qtype + 4].split()[2].split(";")[0])
                if "C6" in text[i + qtype + 5]:
                    lens[10] = float(text[i + qtype + 5].split()[2].split(";")[0])
                if "C7" in text[i + qtype + 6]:
                    lens[11] = float(text[i + qtype + 6].split()[2].split(";")[0])
                if "C8" in text[i + qtype + 7]:
                    lens[12] = float(text[i + qtype + 7].split()[2].split(";")[0])
                if "C9" in text[i + qtype + 8]:
                    lens[13] = float(text[i + qtype + 8].split()[2].split(";")[0])
                if "C10" in text[i + qtype + 9]:
                    lens[14] = float(text[i + qtype + 9].split()[2].split(";")[0])
                if "C11" in text[i + qtype + 10]:
                    lens[15] = float(text[i + qtype + 10].split()[2].split(";")[0])
                if "C12" in text[i + qtype + 11]:
                    lens[16] = float(text[i + qtype + 11].split()[2].split(";")[0])
                if "C13" in text[i + qtype + 12]:
                    lens[17] = float(text[i + qtype + 12].split()[2].split(";")[0])
                if "C14" in text[i + qtype + 13]:
                    lens[18] = float(text[i + qtype + 13].split()[2].split(";")[0])
                if "C15" in text[i + qtype + 14]:
                    lens[19] = float(text[i + qtype + 14].split()[2].split(";")[0])
                if "C16" in text[i + qtype + 15]:
                    lens[20] = float(text[i + qtype + 15].split()[2].split(";")[0])
                if "C17" in text[i + qtype + 16]:
                    lens[21] = float(text[i + qtype + 16].split()[2].split(";")[0])
            lens_data.append(lens)


def convert_qcn_asp():
    lens_asp_data = []
    CF = Conversion_Factor()
    for i in range(0, len(lens_data)):
        lens_asp = [0] * 22
        lens_asp[0] = lens_data[i][0]
        lens_asp[1] = lens_data[i][1]
        lens_asp[2] = lens_data[i][2]
        lens_asp[3] = lens_data[i][3]
        lens_asp[4] = lens_data[i][4]
        lens_asp[5] = lens_data[i][5]
        if lens_asp[0] == "Q-Type":
            lens_asp[6] = lens_data[i][6]
            lens_asp[7] = lens_data[i][7]
            lens_asp[8] = lens_data[i][8] * CF[0][0] + \
                          lens_data[i][9] * CF[1][0] + \
                          lens_data[i][10] * CF[2][0] + \
                          lens_data[i][11] * CF[3][0] + \
                          lens_data[i][12] * CF[4][0] + \
                          lens_data[i][13] * CF[5][0] + \
                          lens_data[i][14] * CF[6][0] + \
                          lens_data[i][15] * CF[7][0] + \
                          lens_data[i][16] * CF[8][0] + \
                          lens_data[i][17] * CF[9][0] + \
                          lens_data[i][18] * CF[10][0] + \
                          lens_data[i][19] * CF[11][0] + \
                          lens_data[i][20] * CF[12][0] + \
                          lens_data[i][21] * CF[13][0]

            lens_asp[9] = lens_data[i][9] * CF[1][1] + \
                          lens_data[i][10] * CF[2][1] + \
                          lens_data[i][11] * CF[3][1] + \
                          lens_data[i][12] * CF[4][1] + \
                          lens_data[i][13] * CF[5][1] + \
                          lens_data[i][14] * CF[6][1] + \
                          lens_data[i][15] * CF[7][1] + \
                          lens_data[i][16] * CF[8][1] + \
                          lens_data[i][17] * CF[9][1] + \
                          lens_data[i][18] * CF[10][1] + \
                          lens_data[i][19] * CF[11][1] + \
                          lens_data[i][20] * CF[12][1] + \
                          lens_data[i][21] * CF[13][1]

            lens_asp[10] = lens_data[i][10] * CF[2][2] + \
                           lens_data[i][11] * CF[3][2] + \
                           lens_data[i][12] * CF[4][2] + \
                           lens_data[i][13] * CF[5][2] + \
                           lens_data[i][14] * CF[6][2] + \
                           lens_data[i][15] * CF[7][2] + \
                           lens_data[i][16] * CF[8][2] + \
                           lens_data[i][17] * CF[9][2] + \
                           lens_data[i][18] * CF[10][2] + \
                           lens_data[i][19] * CF[11][2] + \
                           lens_data[i][20] * CF[12][2] + \
                           lens_data[i][21] * CF[13][2]

            lens_asp[11] = lens_data[i][11] * CF[3][3] + \
                           lens_data[i][12] * CF[4][3] + \
                           lens_data[i][13] * CF[5][3] + \
                           lens_data[i][14] * CF[6][3] + \
                           lens_data[i][15] * CF[7][3] + \
                           lens_data[i][16] * CF[8][3] + \
                           lens_data[i][17] * CF[9][3] + \
                           lens_data[i][18] * CF[10][3] + \
                           lens_data[i][19] * CF[11][3] + \
                           lens_data[i][20] * CF[12][3] + \
                           lens_data[i][21] * CF[13][3]

            lens_asp[12] = lens_data[i][12] * CF[4][4] + \
                           lens_data[i][13] * CF[5][4] + \
                           lens_data[i][14] * CF[6][4] + \
                           lens_data[i][15] * CF[7][4] + \
                           lens_data[i][16] * CF[8][4] + \
                           lens_data[i][17] * CF[9][4] + \
                           lens_data[i][18] * CF[10][4] + \
                           lens_data[i][19] * CF[11][4] + \
                           lens_data[i][20] * CF[12][4] + \
                           lens_data[i][21] * CF[13][4]

            lens_asp[13] = lens_data[i][13] * CF[5][5] + \
                           lens_data[i][14] * CF[6][5] + \
                           lens_data[i][15] * CF[7][5] + \
                           lens_data[i][16] * CF[8][5] + \
                           lens_data[i][17] * CF[9][5] + \
                           lens_data[i][18] * CF[10][5] + \
                           lens_data[i][19] * CF[11][5] + \
                           lens_data[i][20] * CF[12][5] + \
                           lens_data[i][21] * CF[13][5]

            lens_asp[14] = lens_data[i][14] * CF[6][6] + \
                           lens_data[i][15] * CF[7][6] + \
                           lens_data[i][16] * CF[8][6] + \
                           lens_data[i][17] * CF[9][6] + \
                           lens_data[i][18] * CF[10][6] + \
                           lens_data[i][19] * CF[11][6] + \
                           lens_data[i][20] * CF[12][6] + \
                           lens_data[i][21] * CF[13][6]

            lens_asp[15] = lens_data[i][15] * CF[7][7] + \
                           lens_data[i][16] * CF[8][7] + \
                           lens_data[i][17] * CF[9][7] + \
                           lens_data[i][18] * CF[10][7] + \
                           lens_data[i][19] * CF[11][7] + \
                           lens_data[i][20] * CF[12][7] + \
                           lens_data[i][21] * CF[13][7]

            lens_asp[16] = lens_data[i][16] * CF[8][8] + \
                           lens_data[i][17] * CF[9][8] + \
                           lens_data[i][18] * CF[10][8] + \
                           lens_data[i][19] * CF[11][8] + \
                           lens_data[i][20] * CF[12][8] + \
                           lens_data[i][21] * CF[13][8]

            lens_asp[17] = lens_data[i][17] * CF[9][9] + \
                           lens_data[i][18] * CF[10][9] + \
                           lens_data[i][19] * CF[11][9] + \
                           lens_data[i][20] * CF[12][9] + \
                           lens_data[i][21] * CF[13][9]

            lens_asp[18] = lens_data[i][18] * CF[10][10] + \
                           lens_data[i][19] * CF[11][10] + \
                           lens_data[i][20] * CF[12][10] + \
                           lens_data[i][21] * CF[13][10]

            lens_asp[19] = lens_data[i][19] * CF[11][11] + \
                           lens_data[i][20] * CF[12][11] + \
                           lens_data[i][21] * CF[13][11]

            lens_asp[20] = lens_data[i][20] * CF[12][12] + \
                           lens_data[i][21] * CF[13][12]

            lens_asp[21] = lens_data[i][21] * CF[13][13]
        else:
            for n in range(6, 22):
                lens_asp[n] = lens_data[i][n]
        lens_asp_data.append(lens_asp)
    return lens_asp_data


wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "Surface #"
ws.cell(row=1, column=2).value = "Surface Type"
ws.cell(row=1, column=3).value = "Surface Name"
ws.cell(row=1, column=4).value = "Radius"
ws.cell(row=1, column=5).value = "Thickness"
ws.cell(row=1, column=6).value = "Glass"
ws.cell(row=1, column=7).value = "K"
ws.cell(row=1, column=8).value = "Normalization Radius"
ws.cell(row=1, column=9).value = "A2"
ws.cell(row=1, column=10).value = "A4"
ws.cell(row=1, column=11).value = "A6"
ws.cell(row=1, column=12).value = "A8"
ws.cell(row=1, column=13).value = "A10"
ws.cell(row=1, column=14).value = "A12"
ws.cell(row=1, column=15).value = "A14"
ws.cell(row=1, column=16).value = "A16"
ws.cell(row=1, column=17).value = "A18"
ws.cell(row=1, column=18).value = "A20"
ws.cell(row=1, column=19).value = "A22"
ws.cell(row=1, column=20).value = "A24"
ws.cell(row=1, column=21).value = "A26"
ws.cell(row=1, column=22).value = "A28"
ws.cell(row=1, column=23).value = "A30"

data = convert_qcn_asp()
# data = lens_data
for i in range(0, len(data)):
    ws.cell(row=i + 2, column=1).value = i
    for j in range(0, len(data[i])):
        ws.cell(row=i + 2, column=j + 2).value = data[i][j]
wb.save('sample.xlsx')
# print(lens_data)
