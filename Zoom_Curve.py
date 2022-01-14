import win32com
from win32com.client import Dispatch
from openpyxl import load_workbook

# cv = win32com.client.Dispatch('CODEV.Command.111')
# cv.SetStartingDirectory("C:\\CVUSER")
# cv.StartCodeV()
# cv.command('res "C:\\Users\\Alfie\\Desktop\\AA\\new_motor_20210915.len"')
# cv.command('out t tt.txt')
# print(f'OC:{cv.command("eva (Y R1 SI W3 F1 Z1)").split("Command End")[0].split("=")[1]}')
# print('done')

wb = load_workbook('cz.xlsx')
sheet = wb['工作表1']
for i in range(1,202):
    FNO=sheet.cell(i, 1).value
    EFL=sheet.cell(i, 2).value
    cv = win32com.client.Dispatch('CODEV.Command.111')
    cv.SetStartingDirectory("C:\\CVUSER")
    cv.StartCodeV()
    cv.command('res "C:\\Users\\Alfie\\Desktop\\test\\2G_ALL_V4.len"')
    cv.command(f'fno {FNO}')
    cv.command('AUT')
    cv.command('EFP No')
    cv.command('DEL 0.22')
    cv.command(f'EFL Z1 ={EFL}')
    cv.command(f'OAL S1..I Z1 = 38')
    cv.command(f'GO')
    thi_s6=(cv.command(f'wri (thi s2) (thi s10) (thi s19)')).split('Command')[0]
    # thi_s16=cv.command(f'(thi s16)')
    # thi_s20=cv.command(f'(thi s20)')
    print(thi_s6)
    sheet.cell(i, 3).value=thi_s6
    cv.StopCodeV()
    del cv
wb.save('cz.xlsx')
